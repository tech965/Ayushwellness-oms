"""Generic `/payments` endpoints (`app.api.v1.endpoints.payments`) —
list filters (provider/status/q/date range/payment_method), the
order_id-only backward-compatibility contract, eager-loaded
order/customer denormalization, the payment detail view (transactions,
failure reason), and xlsx export. Covers both a Cashfree-style payment
and a plain COD payment (this endpoint is provider-agnostic — see
`app.services.payment_service`'s docstring), so a Cashfree-only
assumption can never silently creep back in.
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from app.core.config import settings
from app.integrations.registry import clear_adapters
from app.models.enums import PaymentStatus, PaymentType
from app.models.mixins import SourceSystem
from app.repositories.payment import PaymentRepository, PaymentTransactionRepository
from app.services.order_service import OrderService
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

pytestmark = pytest.mark.asyncio


@pytest.fixture(autouse=True)
def _configure_cashfree(monkeypatch: pytest.MonkeyPatch):
    # Not exercised over the network in this file -- only here so
    # `Payment.provider == "cashfree"` rows look like a real Cashfree
    # integration would produce them, matching `test_cashfree_payments.py`.
    monkeypatch.setattr(settings, "CASHFREE_CLIENT_ID", "test-client-id")
    monkeypatch.setattr(settings, "CASHFREE_CLIENT_SECRET", "test-client-secret")
    yield
    clear_adapters()


async def _make_order(
    session: AsyncSession,
    *,
    order_number: str,
    customer_name: str = "Asha Test",
    customer_phone: str = "9999999999",
    customer_email: str = "asha@example.com",
) -> uuid.UUID:
    from app.services.customer_service import CustomerService

    customer, _ = await CustomerService(session).upsert_synced_customer(
        source_system=SourceSystem.CASHFREE,
        external_id=f"cust-{order_number}",
        first_name=customer_name.split(" ")[0],
        last_name=" ".join(customer_name.split(" ")[1:]) or None,
        full_name=customer_name,
        email=customer_email,
        phone=customer_phone,
    )
    order = await OrderService(session).create_order(
        actor=None,
        order_number=order_number,
        customer_id=customer.id,
        order_datetime=datetime.now(UTC),
        currency="INR",
        payment_type=PaymentType.PREPAID,
        shipping_charge=0,
        notes=None,
        items=[],
    )
    await session.commit()
    return order.id


async def _make_payment(
    session: AsyncSession,
    *,
    order_id: uuid.UUID,
    status: PaymentStatus = PaymentStatus.PENDING,
    provider: str | None = "cashfree",
    external_id: str | None = None,
    payment_method: str | None = None,
    payment_session_id: str | None = "session_abc",
    amount: Decimal = Decimal("500.00"),
    created_at: datetime | None = None,
):
    metadata: dict = {}
    if payment_method:
        metadata["payment_method"] = payment_method
    if payment_session_id:
        metadata["payment_session_id"] = payment_session_id

    fields: dict = {
        "order_id": order_id,
        "payment_type": PaymentType.PREPAID,
        "status": status,
        "amount": amount,
        "currency": "INR",
        "provider": provider,
        "source_system": SourceSystem.CASHFREE if provider == "cashfree" else None,
        "external_id": external_id,
        "payment_metadata": metadata or None,
        "paid_at": datetime.now(UTC) if status == PaymentStatus.PAID else None,
    }
    if created_at is not None:
        fields["created_at"] = created_at

    payment = await PaymentRepository(session).create(**fields)
    await session.commit()
    return payment


# --- A. Filters, backward compatibility ------------------------------


async def test_order_id_only_filter_stays_backward_compatible(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    """`order_id` alone is the original (Phase 1) filter contract — every
    row `/payments?order_id=` returns must belong to that order, exactly
    as before. (`OrderService.create_order` already creates one baseline
    "manual" `Payment` per order — see that method — so a Cashfree
    checkout attempt added on top makes two rows for the same order, not
    one; this asserts the *filter*, not an exact count that would be
    coupled to that unrelated Phase 1 behavior.)
    """
    order_1 = await _make_order(db_session, order_number="#PAY1")
    order_2 = await _make_order(db_session, order_number="#PAY2")
    await _make_payment(db_session, order_id=order_1, external_id="cf_order_1")
    await _make_payment(db_session, order_id=order_2, external_id="cf_order_2")

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments", params={"order_id": str(order_1)})

    assert response.status_code == 200
    data = response.json()["data"]
    assert len(data) >= 1
    assert all(row["order_id"] == str(order_1) for row in data)
    assert any(row["external_id"] == "cf_order_1" for row in data)


async def test_list_payments_with_no_filters_returns_every_payment(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    order_1 = await _make_order(db_session, order_number="#PAY3")
    order_2 = await _make_order(db_session, order_number="#PAY4")
    await _make_payment(db_session, order_id=order_1, external_id="cf_pay3")
    await _make_payment(db_session, order_id=order_2, external_id="cf_pay4")

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments")

    assert response.status_code == 200
    data = response.json()["data"]
    # 2 orders x (1 auto-created "manual" baseline payment + 1 Cashfree
    # attempt each added above) = 4 -- see the docstring above.
    assert response.json()["meta"]["total_items"] == 4
    assert {row["external_id"] for row in data} >= {"cf_pay3", "cf_pay4"}


async def test_filter_by_provider(db_session: AsyncSession, make_authenticated_client) -> None:
    cashfree_order = await _make_order(db_session, order_number="#PAY5")
    cod_order = await _make_order(db_session, order_number="#PAY6")
    await _make_payment(db_session, order_id=cashfree_order, provider="cashfree")
    await _make_payment(db_session, order_id=cod_order, provider=None)

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments", params={"provider": "cashfree"})

    data = response.json()["data"]
    assert len(data) == 1
    assert data[0]["provider"] == "cashfree"


async def test_filter_by_status(db_session: AsyncSession, make_authenticated_client) -> None:
    order = await _make_order(db_session, order_number="#PAY7")
    await _make_payment(db_session, order_id=order, status=PaymentStatus.PAID, external_id="cf_p")
    await _make_payment(
        db_session,
        order_id=await _make_order(db_session, order_number="#PAY8"),
        status=PaymentStatus.FAILED,
        external_id="cf_f",
    )

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments", params={"status": "paid"})

    data = response.json()["data"]
    assert len(data) == 1
    assert data[0]["status"] == "paid"


async def test_filter_by_payment_method(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    upi_order = await _make_order(db_session, order_number="#PAY9")
    card_order = await _make_order(db_session, order_number="#PAY10")
    await _make_payment(db_session, order_id=upi_order, payment_method="upi", external_id="cf_u")
    await _make_payment(
        db_session, order_id=card_order, payment_method="card", external_id="cf_c"
    )

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments", params={"payment_method": "upi"})

    data = response.json()["data"]
    assert len(data) == 1
    assert data[0]["payment_method"] == "upi"


async def test_filter_by_date_range(db_session: AsyncSession, make_authenticated_client) -> None:
    order = await _make_order(db_session, order_number="#PAY11")
    old_order = await _make_order(db_session, order_number="#PAY12")
    await _make_payment(
        db_session, order_id=order, external_id="cf_recent", created_at=datetime.now(UTC)
    )
    await _make_payment(
        db_session,
        order_id=old_order,
        external_id="cf_old",
        created_at=datetime.now(UTC) - timedelta(days=30),
    )

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        # Scoped to `provider=cashfree` -- the old order's own
        # auto-created "manual" baseline payment (see
        # `OrderService.create_order`) is created *now*, in this test
        # run, same as everything else; only the Cashfree-provider rows'
        # `created_at` was deliberately backdated above.
        response = await client.get(
            "/api/v1/payments",
            params={
                "provider": "cashfree",
                "date_from": (datetime.now(UTC) - timedelta(days=1)).isoformat(),
            },
        )

    data = response.json()["data"]
    assert len(data) == 1
    assert data[0]["external_id"] == "cf_recent"


async def test_search_matches_order_number_customer_and_gateway_order_id(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    target_order = await _make_order(
        db_session,
        order_number="#AWL55001",
        customer_name="Ravi Kumar",
        customer_phone="8888800001",
    )
    other_order = await _make_order(
        db_session, order_number="#AWL55002", customer_name="Someone Else"
    )
    await _make_payment(db_session, order_id=target_order, external_id="cf_target")
    await _make_payment(db_session, order_id=other_order, external_id="cf_other")

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        by_order_number = await client.get("/api/v1/payments", params={"q": "AWL55001"})
        by_customer_name = await client.get("/api/v1/payments", params={"q": "Ravi Kumar"})
        by_gateway_id = await client.get("/api/v1/payments", params={"q": "cf_target"})

    # Order-number/customer-name search also matches that order's own
    # auto-created "manual" baseline payment (see
    # `OrderService.create_order`) -- both rows share the same order and
    # customer. The gateway-id search is exact to the Cashfree row alone.
    for response in (by_order_number, by_customer_name):
        data = response.json()["data"]
        assert len(data) >= 1
        assert any(row["external_id"] == "cf_target" for row in data)
        assert all(row["order_number"] == "#AWL55001" for row in data)

    gateway_data = by_gateway_id.json()["data"]
    assert len(gateway_data) == 1
    assert gateway_data[0]["external_id"] == "cf_target"


async def test_invalid_status_filter_returns_422(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments", params={"status": "not-a-real-status"})
    assert response.status_code == 422


# --- B. Denormalized order/customer fields, no N+1 -------------------


async def test_list_response_includes_denormalized_order_and_customer_fields(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    order_id = await _make_order(
        db_session,
        order_number="#AWL66001",
        customer_name="Priya Sharma",
        customer_phone="7777700001",
        customer_email="priya@example.com",
    )
    await _make_payment(
        db_session,
        order_id=order_id,
        external_id="cf_priya",
        payment_method="netbanking",
        status=PaymentStatus.PAID,
    )

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments", params={"provider": "cashfree"})

    row = response.json()["data"][0]
    assert row["order_number"] == "#AWL66001"
    assert row["customer_name"] == "Priya Sharma"
    assert row["customer_phone"] == "7777700001"
    assert row["customer_email"] == "priya@example.com"
    assert row["external_id"] == "cf_priya"
    assert row["payment_method"] == "netbanking"
    assert row["payment_session_id"] == "session_abc"


# --- C. Detail view: transactions, failure reason ---------------------


async def test_payment_detail_includes_transactions_and_failure_reason(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    order_id = await _make_order(db_session, order_number="#AWL77001")
    payment = await _make_payment(
        db_session, order_id=order_id, external_id="cf_fail", status=PaymentStatus.FAILED
    )
    await PaymentTransactionRepository(db_session).create(
        payment_id=payment.id,
        gateway="cashfree",
        gateway_transaction_id="cfpay_1",
        status=PaymentStatus.FAILED,
        amount=payment.amount,
        raw_payload={
            "type": "PAYMENT_FAILED_WEBHOOK",
            "payment": {"payment_method": "upi"},
            "error_details": {
                "error_code": "insufficient_funds",
                "error_description": "The customer's account has insufficient funds.",
            },
        },
    )
    await db_session.commit()

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get(f"/api/v1/payments/{payment.id}")

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["status"] == "failed"
    assert len(body["transactions"]) == 1
    txn = body["transactions"][0]
    assert txn["gateway_transaction_id"] == "cfpay_1"
    assert txn["event_type"] == "PAYMENT_FAILED_WEBHOOK"
    assert txn["payment_method"] == "upi"
    assert txn["error_reason"] == "The customer's account has insufficient funds."


async def test_payment_detail_unknown_id_returns_404(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get(f"/api/v1/payments/{uuid.uuid4()}")
    assert response.status_code == 404


async def test_list_and_detail_require_permission(
    db_session: AsyncSession, client: AsyncClient
) -> None:
    order_id = await _make_order(db_session, order_number="#AWL88001")
    payment = await _make_payment(db_session, order_id=order_id, external_id="cf_noauth")

    list_response = await client.get("/api/v1/payments")
    detail_response = await client.get(f"/api/v1/payments/{payment.id}")
    assert list_response.status_code == 401
    assert detail_response.status_code == 401


# --- D. Export ----------------------------------------------------------


async def test_export_payments_streams_a_real_xlsx_workbook(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    order_id = await _make_order(db_session, order_number="#AWL99001", customer_name="Export Test")
    await _make_payment(
        db_session,
        order_id=order_id,
        external_id="cf_export",
        payment_method="upi",
        status=PaymentStatus.PAID,
        amount=Decimal("999.00"),
    )

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        response = await client.get("/api/v1/payments/export")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    sheet = workbook.active
    header_row = [cell.value for cell in sheet[1]]
    assert "Order Number" in header_row
    assert "Cashfree Order ID" in header_row
    # >=2 data rows: the order's own auto-created "manual" baseline
    # payment (see `OrderService.create_order`) plus the Cashfree one
    # added above -- find the Cashfree row rather than assuming a
    # position among them.
    data_rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    matching = [row for row in data_rows if "cf_export" in row]
    assert len(matching) == 1
    assert "#AWL99001" in matching[0]
    assert "Export Test" in matching[0]


async def test_export_payments_requires_permission(
    db_session: AsyncSession, client: AsyncClient
) -> None:
    response = await client.get("/api/v1/payments/export")
    assert response.status_code == 401


# --- E. Confirms this endpoint is not Cashfree-exclusive ---------------


async def test_a_cod_payment_with_no_provider_still_lists_and_shows_detail(
    db_session: AsyncSession, make_authenticated_client
) -> None:
    """`/payments` predates Cashfree (Phase 1: `OrderService.create_order`
    already created `Payment` rows for COD/manual payments) — a payment
    with no gateway provider must keep working exactly as it did before.
    """
    order_id = await _make_order(db_session, order_number="#AWL10001")
    payment = await _make_payment(db_session, order_id=order_id, provider=None, external_id=None)

    async with await make_authenticated_client(
        db_session, permission_codes=["payments.read"]
    ) as client:
        listed = await client.get("/api/v1/payments")
        detail = await client.get(f"/api/v1/payments/{payment.id}")

    assert listed.status_code == 200
    assert detail.status_code == 200
    row = next(r for r in listed.json()["data"] if r["id"] == str(payment.id))
    assert row["provider"] is None
    assert row["external_id"] is None
    assert detail.json()["data"]["transactions"] == []
