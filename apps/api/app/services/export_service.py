"""Renders already-fetched Order/Payment rows as a real `.xlsx` workbook.

Deliberately takes fully-loaded ORM instances (with the relationships the
caller's `search_query`/`list_for_export` already eager-loaded) rather
than touching the session itself, so this stays a pure, easily-testable
formatting layer.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.order import Order
from app.models.payment import Payment


class ExportService:
    # Safety cap so an unfiltered export can't try to stream the entire
    # orders table into memory in one request.
    MAX_ROWS = 10_000

    HEADERS = [
        "Order Number",
        "Order Date",
        "Customer Name",
        "Phone",
        "Email",
        "Products",
        "Total Quantity",
        "Amount",
        "Currency",
        "Payment Type",
        "Payment Status",
        "Order Status",
        "Fulfillment Status",
        "Shipment Status",
        "Courier",
        "AWB",
        "Created At",
    ]

    def orders_to_xlsx(self, orders: list[Order]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Orders"
        sheet.append(self.HEADERS)

        for order in orders:
            # An order can technically have more than one Shipment record
            # (a re-ship after RTO, for instance) — the export shows the
            # most recently created one, matching what the Order Detail
            # page treats as "the" shipment.
            shipment = order.shipments[-1] if order.shipments else None
            products = "; ".join(f"{item.product_name} x{item.quantity}" for item in order.items)
            total_quantity = sum(item.quantity for item in order.items)

            sheet.append(
                [
                    order.order_number,
                    _naive(order.order_datetime),
                    order.customer.full_name if order.customer else None,
                    order.customer.phone if order.customer else None,
                    order.customer.email if order.customer else None,
                    products,
                    total_quantity,
                    float(order.total_amount),
                    order.currency,
                    order.payment_type.value,
                    order.payment_status.value,
                    order.status.value,
                    order.fulfillment_status.value,
                    shipment.current_status.value if shipment else None,
                    shipment.courier.name if shipment and shipment.courier else None,
                    shipment.awb if shipment else None,
                    _naive(order.created_at),
                ]
            )

        for index in range(1, len(self.HEADERS) + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 22

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    PAYMENT_HEADERS = [
        "Order Number",
        "Customer Name",
        "Phone",
        "Email",
        "Amount",
        "Currency",
        "Provider",
        "Payment Method",
        "Status",
        "Cashfree Order ID",
        "Gateway Transaction ID",
        "Created At",
        "Paid At",
    ]

    def payments_to_xlsx(self, payments: list[Payment]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Payments"
        sheet.append(self.PAYMENT_HEADERS)

        for payment in payments:
            order = payment.order
            customer = order.customer if order is not None else None
            metadata = payment.payment_metadata or {}

            sheet.append(
                [
                    order.order_number if order is not None else None,
                    customer.full_name if customer is not None else None,
                    customer.phone if customer is not None else None,
                    customer.email if customer is not None else None,
                    float(payment.amount),
                    payment.currency,
                    payment.provider,
                    metadata.get("payment_method"),
                    payment.status.value,
                    payment.external_id,
                    payment.external_transaction_id,
                    _naive(payment.created_at),
                    _naive(payment.paid_at),
                ]
            )

        for index in range(1, len(self.PAYMENT_HEADERS) + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 22

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


def _naive(value):  # noqa: ANN001, ANN202
    """openpyxl rejects timezone-aware datetimes outright — every
    timestamp in this codebase is stored/returned as timezone-aware UTC
    (`AwareDateTime`), so every one written to a cell needs this.
    """
    return value.replace(tzinfo=None) if value is not None else None
